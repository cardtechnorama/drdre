"""Logging and optional Excel export for ranked candidate DataFrames."""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Sequence

import pandas as pd

_DECIMAL_FORMATS: dict[str, str] = {
    "Candidate": "0",
    "Rank overall": "0",
    "Avg FEA rank": "0.000",
    "Final avg rank": "0.000",
    "rmse_MM": "0.0000",
}


def _number_format_for(col: str) -> str:
    if col in _DECIMAL_FORMATS:
        return _DECIMAL_FORMATS[col]
    if col.startswith("Rank "):
        return "0"
    low = col.lower()
    if "%" in col or "color score" in low:
        return "0.00"
    if "stress" in low:
        return "0.000"
    if any(k in low for k in ("motion", "interface", "disp")):
        return "0.000000"
    return "0.0000"


def _round_for_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        name = str(col)
        if name == "Candidate" or name == "Rank overall" or name.startswith("Rank "):
            continue
        fmt = _number_format_for(name)
        digits = max(0, len(fmt.split(".")[1])) if "." in fmt else 0
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].round(digits)
    return out


def write_excel(df: pd.DataFrame, out_xlsx: Path, sheet: str = "ranked") -> Path:
    """Write the ranked DataFrame to xlsx with per-column Excel number formats."""
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    display = _round_for_display(df)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        display.to_excel(writer, index=False, sheet_name=sheet)

    from openpyxl import load_workbook

    wb = load_workbook(out_xlsx)
    ws = wb[sheet]
    for idx, cell in enumerate(ws[1], start=1):
        name = cell.value
        if name is None:
            continue
        fmt = _number_format_for(str(name))
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=idx).number_format = fmt
    wb.save(out_xlsx)
    return out_xlsx


_DEFAULT_LOG_COLUMNS: tuple[str, ...] = (
    "Rank overall",
    "Candidate",
    "rmse_MM",
    "Color score (Safe+Caution)",
    "Avg FEA rank",
    "Final avg rank",
)


def log_top_candidates(
    df: pd.DataFrame,
    top_n: int = 10,
    columns: Sequence[str] | None = None,
    log: Callable[[str], None] = print,
) -> pd.DataFrame:
    """Log the top-N candidates by ``Final avg rank`` and return the slice."""
    cols = [c for c in (columns or _DEFAULT_LOG_COLUMNS) if c in df.columns]
    if not cols:
        log("No ranking columns available to report.")
        return df.head(0)

    n = min(int(top_n), len(df))
    view = df[cols].head(n).copy()
    for c in view.columns:
        if c in {"Candidate", "Rank overall"} or str(c).startswith("Rank "):
            continue
        if pd.api.types.is_numeric_dtype(view[c]):
            view[c] = view[c].round(4)

    log("")
    log(f"Top {n} candidates by Final avg rank:")
    log(view.to_string(index=False))
    return view
